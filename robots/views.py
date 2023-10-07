import json
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.http import JsonResponse
from django.views.generic.edit import CreateView
from .models import Robot
from .validators import validate_json_data
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from django.views import View
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
from django.http import FileResponse


class RobotCreateView(CreateView):
    """
        Класс принимающий и обрабатывающий информацию в формате JSON
    """
    model = Robot

    def post(self, request, *args, **kwargs):
        try:
            uploaded_file = json.loads(request.body.decode('utf-8'))
            validation_errors = []
            for data in uploaded_file:
                try:
                    validate_json_data(data)
                    model = data['model']
                    version = data['version']
                    created = data['created']

                    robots_to_insert = [
                        Robot(model=model, version=version, created=created)
                    ]
                    Robot.objects.bulk_create(robots_to_insert)
                except ValidationError as e:
                    validation_errors.append(str(e))
            if validation_errors:
                return JsonResponse({'errors': validation_errors}, status=400)
            return JsonResponse({'message': 'Data received and processed successfully'})
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)


class UploadExelView(View):
    """
     Класс для вигрузки exel файла
    """
    wb = Workbook()

    def excel_data(self):
        """
            Детализация по версии в Exel sheets
        """
        one_week_ago = timezone.now() - timedelta(days=7)
        data = Robot.objects.filter(created__gte=one_week_ago).values('model', 'version').annotate(
            count=Count('version'))

        for entry in data:

            model = entry['model']
            if model in self.wb.sheetnames:
                self.wb.active = self.wb.sheetnames.index(model)
                ws = self.wb.active
            else:
                ws = self.wb.create_sheet(model)
            if len(ws._cells) <= 1:
                ws.append(['Модель', 'Версия', 'Количество за неделю'])

            ws.append([entry['model'], entry['version'], entry['count']])

        xlsx_path = 'detailed.xlsx'
        self.wb.save(xlsx_path)
        return xlsx_path

    def get(self, request):
        if request.method == 'GET':
            xlsx_path = self.excel_data()
            response = FileResponse(open(xlsx_path, 'rb'),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="detailed.xlsx"'
            return response
